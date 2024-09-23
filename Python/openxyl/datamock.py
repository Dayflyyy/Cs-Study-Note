import openpyxl
from openpyxl.utils import get_column_letter
import random

# 创建一个新的工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 设置表头
headers = ["商品名称","商品条码","商品规格","数量单位","商品数量","商品单价","商品总价","使用小单..."]
ws.append(headers)

# 生成模拟数据
for i in range(1, 100):
    row = [

        f"商品{i}",  # 商品名称
        f"123456789{i}",  # 商品条码
        f"规格{i}",  # 商品规格
        "件",  # 数量单位
        random.randint(1, 100),  # 商品数量
        round(random.uniform(10, 100), 2),  # 商品单价
        round(random.uniform(1000, 10000), 2),  # 商品总价
    random.choice(["是", "否"])  # 使用小单
    ]
    ws.append(row)

# 自动调整列宽
for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    max_length = max(len(str(cell.value)) for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# 保存工作簿
wb.save("模拟数据商品.xlsx")